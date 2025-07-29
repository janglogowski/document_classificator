import os
import time
import uuid
import shutil
from datetime import datetime

import yaml
import cv2
import joblib
import pytesseract
import numpy as np
import torch
import torch.nn as nn
from torchvision import transforms, models
from PIL import Image
from pdf2image import convert_from_path
from collections import OrderedDict

#--- load config ---#
cfg  = yaml.safe_load(open("config.yaml", encoding="utf-8"))
ROOT = os.path.abspath(cfg["paths"]["project_root"])

#--- manual override ---#
ocr_engine     = "easy_ocr"         # "easy_ocr"/"tesseract_ocr"
vsd_classifier = "cnn"              # "cnn"/"tfidf_lr"
doc_classifier = "tfidf_lr"         # "cnn"/"tfidf_lr"
level          = "level3"           # "level1"/"level2"/"level3"

#--- paths ---#
P       = cfg["paths"]
DATA    = P["data"]
MODELS  = DATA["models"]
TESTS   = P["tests"]
POPPLER = P["poppler"]["bin_path"]

INPUT_FOLDER  = os.path.join(ROOT, TESTS["input"])
OUTPUT_FOLDER = os.path.join(ROOT, TESTS["output"])
METADATA_CSV  = os.path.join(ROOT, TESTS["metadata"])
os.makedirs(INPUT_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

#--- OCR settings ---# 
ocr_cfg  = cfg[ocr_engine]
img_exts = tuple(ocr_cfg["image_extensions"])
pdf_ext  = ocr_cfg["pdf_extension"]
pdf_dpi  = ocr_cfg["pdf_dpi"]
max_dim  = ocr_cfg["max_dim"]
lang     = ocr_cfg["lang"]

#--- tesseract OCR function ---#
def ocr_tesseract(path):
    ext = os.path.splitext(path)[1].lower()
    text = ""
    if ext in img_exts:
        img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
        if img is None: return ""
        h,w = img.shape
        if max(h,w) > max_dim:
            s = max_dim / max(h,w)
            img = cv2.resize(img, (int(w*s),int(h*s)))
        _,img = cv2.threshold(img,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
        text = pytesseract.image_to_string(img, lang=lang)
    elif ext == pdf_ext:
        pages = convert_from_path(path, dpi=pdf_dpi, poppler_path=POPPLER)
        for pg in pages:
            arr = cv2.cvtColor(np.array(pg), cv2.COLOR_RGB2BGR)
            gray = cv2.cvtColor(arr, cv2.COLOR_BGR2GRAY)
            _,gray = cv2.threshold(gray,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
            text += pytesseract.image_to_string(gray, lang=lang) + "\n"
    return text.strip()

#--- easyOCR function ---#
if ocr_engine == "easy_ocr":
    import easyocr
    langs = [lang] if isinstance(lang,str) else list(lang)
    reader = easyocr.Reader(langs, gpu=False)
    def ocr_easyocr(path):
        ext = os.path.splitext(path)[1].lower()
        if ext in img_exts:
            img = cv2.imread(path)
            if img is None: return ""
            h,w = img.shape[:2]
            if max(h,w)>max_dim:
                s = max_dim/max(h,w)
                img = cv2.resize(img,(int(w*s),int(h*s)))
            return "\n".join(reader.readtext(img,detail=0))
        elif ext == pdf_ext:
            pages = convert_from_path(path, dpi=pdf_dpi, poppler_path=POPPLER)
            out=[]
            for pg in pages:
                arr=np.array(pg)[:,:,::-1]
                out.append("\n".join(reader.readtext(arr,detail=0)))
            return "\n\n".join(out)
        return ""

#--- select OCR ---#
if ocr_engine=="tesseract_ocr":
    pytesseract.pytesseract.tesseract_cmd = cfg["tesseract_ocr"]["tesseract_cmd"]
    ocr_image = ocr_tesseract
else:
    ocr_image = ocr_easyocr

#--- generate filename & log metadata ---#
def generate_filename(doc_type, ext):
    uid = uuid.uuid4().hex[:6]
    safe = doc_type.replace(" ","_").lower()
    return f"{safe}_{uid}{ext}"

#--- log metadata ---#
def log_metadata(doc_type, src, new, classification_time=None):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill

    excel_path = METADATA_CSV.replace(".csv", ".xlsx")
    file_exists = os.path.exists(excel_path)

    headers = ["input_file", "ocr_engine", "vsd_classifier", "doc_classifier", "level",
               "predicted_type", "internal_name", "timestamp", "classification_time_sec"]
    row = [os.path.basename(src), ocr_engine, vsd_classifier, doc_classifier,
           level, doc_type, new, datetime.now().isoformat(), classification_time]

    if file_exists:
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "metadata"
        ws.append(headers)
    ws.append(row)

    predicted = doc_type.lower()
    file_name = os.path.basename(src).split('.')[0].lower()
    if predicted in file_name:
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        last_row = ws.max_row
        for col in range(1, len(headers)+1):
            ws.cell(last_row, col).fill = green_fill

    wb.save(excel_path)

#--- load text‐based models for doc vs drw ---#
if vsd_classifier=="tfidf_lr":
    lr_vs = MODELS["lr_doc_vs_drw"]
    base_vs = os.path.join(ROOT,lr_vs,ocr_engine)
    vsd_model = joblib.load(os.path.join(base_vs,f"lr_doc_vs_drw_{level}.pkl"))
    vsd_vect  = joblib.load(os.path.join(base_vs,f"doc_vs_drw_vect_{level}.pkl"))

if vsd_classifier == "cnn":
    cnn_vs_root = os.path.join(ROOT, MODELS["cnn_doc_vs_drw"])
    state = torch.load(os.path.join(cnn_vs_root, f"cnn_doc_vs_drw_{level}.pth"), map_location="cpu")

    from torchvision.models import mobilenet_v2
    vsd_cnn = mobilenet_v2(weights=None)
    vsd_cnn.classifier[1] = nn.Linear(vsd_cnn.last_channel, 2)

    if isinstance(state, OrderedDict):
        vsd_cnn.load_state_dict(state)

    vsd_cnn.eval()
    vsd_cnn_tf = transforms.Compose([
        transforms.Resize((256, 256)),
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
    ])

    def predict_vsd_cnn(path):
        ext = os.path.splitext(path)[1].lower()
        if ext in img_exts:
            pil = Image.open(path).convert("RGB")
        elif ext == pdf_ext:
            pg = convert_from_path(path, dpi=pdf_dpi, poppler_path=POPPLER)
            pil = pg[0].convert("RGB") if pg else None
        else:
            return "undefined"
        if pil is None:
            return "undefined"
        t = vsd_cnn_tf(pil).unsqueeze(0)
        idx = vsd_cnn(t).argmax(dim=1).item()
        return ["document", "tech_drw"][idx]

#--- load doc‐type classifier ---#
if doc_classifier=="tfidf_lr":
    lr_dt = MODELS["lr_doc_type_classifier"][ocr_engine]
    base_dt = os.path.join(ROOT,lr_dt)
    dt_model = joblib.load(os.path.join(base_dt,f"doc_type_classifier_{level}.pkl"))
    dt_vect  = joblib.load(os.path.join(base_dt,f"doc_type_vect_{level}.pkl"))
else:
    cnn_dt_root = os.path.join(ROOT,MODELS["cnn_doc_type_classifier"])
    state2 = torch.load(os.path.join(cnn_dt_root,f"cnn_doc_type_classifier_{level}.pth"),map_location="cpu")
    dt_cnn = models.resnet18(pretrained=False)
    dt_cnn.fc = nn.Linear(dt_cnn.fc.in_features,6)
    if isinstance(state2,OrderedDict):
        dt_cnn.load_state_dict(state2)
    dt_cnn.eval()
    dt_cnn_tf = transforms.Compose([
        transforms.Resize((256,256)),
        transforms.ToTensor(),
        transforms.Normalize([0.485,0.456,0.406],[0.229,0.224,0.225])])
    
    names = [
        "BOM",
        "DAILY_REPORT",
        "INSPECTION_REPORT",
        "MAINTENANCE_LOG",
        "PRODUCT_DATA_SHEET",
        "QUALITY_CHECKLIST"]
    
    def predict_dt_cnn(path):
        ext=os.path.splitext(path)[1].lower()
        if ext in img_exts:
            pil=Image.open(path).convert("RGB")
        elif ext==pdf_ext:
            pg=convert_from_path(path,dpi=pdf_dpi,poppler_path=POPPLER)
            pil=pg[0].convert("RGB") if pg else None
        else: return "undefined"
        if pil is None: return "undefined"
        t=dt_cnn_tf(pil).unsqueeze(0)
        idx=dt_cnn(t).argmax(dim=1).item()
        return names[idx]

#--- main loop ---#
print(f"\n=== Watching {INPUT_FOLDER} | OCR={ocr_engine} | VSD={vsd_classifier} | DT={doc_classifier} ===\n")
while True:
    for fn in sorted(os.listdir(INPUT_FOLDER)):
        ext = os.path.splitext(fn)[1].lower()
        if ext not in img_exts + (pdf_ext,): continue
        src = os.path.join(INPUT_FOLDER,fn)
        print(f"[→] {fn}")

        start_time = time.time()

        text = ""
        if vsd_classifier == "tfidf_lr" or doc_classifier == "tfidf_lr":
            text = ocr_image(src).strip()

        # stage 1: doc vs drw
        if vsd_classifier == "tfidf_lr":
            p1 = vsd_model.predict(vsd_vect.transform([text]))[0] if text else "document"
        else:
            p1 = predict_vsd_cnn(src)

        if p1 != "document":
            dstype = "tech_drw"
        else:
        # stage 2: doc‐type classifier
            if doc_classifier == "tfidf_lr":
                text = ocr_image(src).strip()
                dstype = dt_model.predict(dt_vect.transform([text]))[0] if text else "undefined"
            else:
                dstype = predict_dt_cnn(src)

        classification_time = round(time.time() - start_time, 3)

        new = generate_filename(dstype, ext)
        shutil.move(src, os.path.join(OUTPUT_FOLDER, new))
        log_metadata(dstype, src, new, classification_time=classification_time)
        print(f"[OK] → {new} (classified in {classification_time}s)\n")

    time.sleep(1.0)
