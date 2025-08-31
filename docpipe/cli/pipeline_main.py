import os
import time
import shutil
from openpyxl import Workbook, load_workbook
import yaml
from docpipe.cli.ocr_utils import get_ocr_function
from docpipe.cli.classification_utils import (load_vsd_model, load_doc_type_model,predict_doc_vs_drw, predict_doc_type)
from docpipe.cli.file_utils import generate_filename, log_metadata

# === Config === #
cfg = yaml.safe_load(open("config.yaml", encoding="utf-8"))
ROOT = os.path.abspath(cfg["paths"]["project_root"])

# === User parameters === #
ocr_engine     = "easy_ocr"         # "easy_ocr"/"tesseract_ocr"
vsd_classifier = "cnn"              # "cnn"/"tfidf_lr"
doc_classifier = "tfidf_lr"         # "cnn"/"tfidf_lr"
level          = "level3"

# === Paths === #
P = cfg["paths"] 
INPUT_FOLDER   = os.path.join(ROOT, P["tests"]["input"])
OUTPUT_FOLDER  = os.path.join(ROOT, P["tests"]["output"])
METADATA_CSV   = os.path.join(ROOT, P["tests"]["metadata"])
BENCHMARK_XLSX = os.path.join(ROOT, "tests", "benchmark_results.xlsx")


os.makedirs(INPUT_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# === Excel init === #
if not os.path.exists(BENCHMARK_XLSX):
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark"
    ws.append([
        "filename", "dstype", "ocr_engine", "vsd_classifier", "doc_classifier", "level",
        "load_time", "ocr_time", "stage1_time", "stage2_time",
        "move_time", "meta_time", "total_time"
    ])
    wb.save(BENCHMARK_XLSX)

def log_to_excel(row):
    wb = load_workbook(BENCHMARK_XLSX)
    ws = wb.active
    ws.append(row)
    wb.save(BENCHMARK_XLSX)

# === Initialize OCR === #
ocr_image = get_ocr_function(ocr_engine, cfg)

# === Load models === #
vsd_model = load_vsd_model(vsd_classifier, ocr_engine, level, cfg)
doc_model, doc_names = load_doc_type_model(doc_classifier, ocr_engine, level, cfg)

print(f"\n=== Watching {INPUT_FOLDER} | OCR={ocr_engine} | VSD={vsd_classifier} | DT={doc_classifier} ===\n")

while True:
    for fn in sorted(os.listdir(INPUT_FOLDER)):
        ext = os.path.splitext(fn)[1].lower()
        if ext not in cfg[ocr_engine]["image_extensions"] + [cfg[ocr_engine]["pdf_extension"]]:
            continue

        src = os.path.join(INPUT_FOLDER, fn)
        print(f"[>] {fn}")

        t0 = time.perf_counter()

        # OCR text if needed
        text = ""
        t_ocr0 = time.perf_counter()
        if vsd_classifier == "tfidf_lr" or doc_classifier == "tfidf_lr":
            text = ocr_image(src).strip()
        t_ocr1 = time.perf_counter()

        # Stage 1
        t_s10 = time.perf_counter()
        pred_vsd = predict_doc_vs_drw(src, vsd_classifier, vsd_model, text, cfg, engine=ocr_engine)
        t_s11 = time.perf_counter()

        # Stage 2
        t_s20 = time.perf_counter()
        if pred_vsd == "document":
            dstype = predict_doc_type(src, doc_classifier, (doc_model, doc_names), ocr_image, text, cfg)
        elif pred_vsd == "tech_drw":
            dstype = "tech_drw"
        else:
            dstype = "UNDEFINED"
        t_s21 = time.perf_counter()

        # Move file
        t_move0 = time.perf_counter()
        new_name = generate_filename(dstype, ext)
        shutil.move(src, os.path.join(OUTPUT_FOLDER, new_name))
        t_move1 = time.perf_counter()

        # Metadata
        t_meta0 = time.perf_counter()
        log_metadata(METADATA_CSV, dstype, src, new_name, ocr_engine, vsd_classifier, doc_classifier, level,
                     round(time.perf_counter() - t0, 3))
        t_meta1 = time.perf_counter()

        # Times
        load_time   = 0.0
        ocr_time    = t_ocr1 - t_ocr0
        stage1_time = t_s11 - t_s10
        stage2_time = t_s21 - t_s20
        move_time   = t_move1 - t_move0
        meta_time   = t_meta1 - t_meta0
        total_time  = time.perf_counter() - t0

        log_to_excel([
            fn, dstype, ocr_engine, vsd_classifier, doc_classifier, level,
            round(load_time,3), round(ocr_time,3), round(stage1_time,3), round(stage2_time,3),
            round(move_time,3), round(meta_time,3), round(total_time,3)
        ])

        print(f"[OK] >>>> {new_name} (total={total_time:.3f}s | ocr={ocr_time:.3f}s | s1={stage1_time:.3f}s | s2={stage2_time:.3f}s)")

    time.sleep(1.0)