import os
import re
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

_CANON_SYNONYMS = {
    "BOM": {
        "bom", "boms", "bill_of_material", "bill_of_materials", "billofmaterials"
    },
    "DAILY_REPORT": {
        "daily_report", "daily_reports", "dailyreport"
    },
    "INSPECTION_REPORT": {
        "inspection_report", "inspection_reports", "inspectionreport"
    },
    "MAINTENANCE_LOG": {
        "maintenance_log", "maintenance_logs", "maintenancelog"
    },
    "PRODUCT_DATA_SHEET": {
        "product_data_sheet", "product_data_sheets", "datasheet", "data_sheet", "pds"
    },
    "QUALITY_CHECKLIST": {
        "quality_checklist", "quality_checklists", "qualitychecklist", "qc"
    },
    "TECH_DRW": {
        "tech_drw", "techdrw", "technical_drawing", "technicaldrawing",
        "tech_drawing", "techdrawing", "drw", "drawing"
    },
}

_CANON_MAP = {
    # BOM
    "bom": "BOM",
    "boms": "BOM",
    "bill_of_material": "BOM",
    "bill_of_materials": "BOM",
    "billofmaterials": "BOM",
    # DAILY_REPORT
    "daily_report": "DAILY_REPORT",
    "daily_reports": "DAILY_REPORT",
    "dailyreport": "DAILY_REPORT",
    # INSPECTION_REPORT
    "inspection_report": "INSPECTION_REPORT",
    "inspection_reports": "INSPECTION_REPORT",
    "inspectionreport": "INSPECTION_REPORT",
    # MAINTENANCE_LOG
    "maintenance_log": "MAINTENANCE_LOG",
    "maintenance_logs": "MAINTENANCE_LOG",
    "maintenancelog": "MAINTENANCE_LOG",
    # PRODUCT_DATA_SHEET
    "product_data_sheet": "PRODUCT_DATA_SHEET",
    "product_data_sheets": "PRODUCT_DATA_SHEET",
    "datasheet": "PRODUCT_DATA_SHEET",
    "data_sheet": "PRODUCT_DATA_SHEET",
    "pds": "PRODUCT_DATA_SHEET",
    # QUALITY_CHECKLIST
    "quality_checklist": "QUALITY_CHECKLIST",
    "quality_checklists": "QUALITY_CHECKLIST",
    "qualitychecklist": "QUALITY_CHECKLIST",
    "qc": "QUALITY_CHECKLIST",
    # TECH_DRW
    "tech_drw": "TECH_DRW",
    "techdrw": "TECH_DRW",
    "technical_drawing": "TECH_DRW",
    "technicaldrawing": "TECH_DRW",
    "tech_drawing": "TECH_DRW",
    "techdrawing": "TECH_DRW",
    "drw": "TECH_DRW",
    "drawing": "TECH_DRW",
}


def _norm_token(s: str) -> str:
    s = (s or "").lower()
    s = s.replace("-", "_")
    s = re.sub(r"[^a-z0-9_]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s

def _canonicalize_doc_type(label: str) -> str:
    """Sprowadza dowolny wariant (np. 'boms', 'technical_drawing') do kanonu."""
    norm = _norm_token(label)
    return _CANON_MAP.get(
        norm,
        _CANON_MAP.get(norm.rstrip("s"), (norm.upper() if norm else "UNDEFINED"))
    )


def _filename_matches_canon(predicted_canon: str, filename_noext: str) -> bool:
    """
    True jeśli nazwa pliku (bez rozszerzenia) jednoznacznie sugeruje przewidziany typ.
    Działa dla nazw z/bez separatorów: 'bom_001', 'bom001', 'projectinspectionreport7', 'tech_drw_12' itd.
    """
    base = _norm_token(filename_noext)      
    base_flat = base.replace("_", "")       
    tokens = set(base.split("_"))           

    syns = _CANON_SYNONYMS.get(predicted_canon, set())
    if not syns:
        return False

    for syn in syns:
        syn_norm = _norm_token(syn)          
        syn_flat = syn_norm.replace("_", "") 

        if syn_norm in tokens:
            return True
        if syn_flat and syn_flat in base_flat:
            return True
        hay = base.replace("_", " ")
        pat = rf"(?<![a-z0-9]){re.escape(syn_norm).replace('_', r'[_\s-]')}(?![a-z0-9])"
        if re.search(pat, hay):
            return True

    return False


def generate_filename(doc_type: str, ext: str) -> str:
    uid = uuid.uuid4().hex[:6]
    safe = _norm_token(doc_type).replace("_", "")
    return f"{safe}_{uid}{ext}"

def _open_or_create_xlsx(xlsx_path: Path):
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb["metadata"] if "metadata" in wb.sheetnames else wb.active
        ws.title = "metadata"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "metadata"
        ws.append([
            "input_file", "ocr_engine", "vsd_classifier", "doc_classifier", "level",
            "predicted_type", "internal_name", "timestamp", "classification_time_sec"
        ])
    return wb, ws

def log_metadata(excel_path, doc_type, src, new_name, ocr, vsd, doc_cls, level, duration):
    xlsx = Path(str(excel_path).replace(".csv", ".xlsx"))
    wb, ws = _open_or_create_xlsx(xlsx)

    canon = _canonicalize_doc_type(doc_type)

    row = [
        os.path.basename(src), ocr, vsd, doc_cls, level,
        canon, new_name, datetime.now().isoformat(), duration
    ]
    ws.append(row)

    base_noext = os.path.splitext(os.path.basename(src))[0]
    if _filename_matches_canon(canon, base_noext):
        last = ws.max_row
        for col in range(1, ws.max_column + 1):
            ws.cell(last, col).fill = _GREEN

    wb.save(xlsx)
